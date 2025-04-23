import cv2
import numpy as np
import csv
import os
from datetime import datetime
import face_recognition
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd

print("OpenCV version:", cv2.__version__)

# Initialize face detection with better parameters
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
if face_cascade.empty():
    print("Error: Could not load face cascade classifier")
    exit()

# Adjust these parameters for even faster performance
TOLERANCE = 0.6
MODEL = "hog"    # Keep 'hog' for faster detection
NUM_JITTERS = 1  # Keep at 1 for speed
FRAME_RESIZE = 0.2  # Reduce to 20% for faster processing
PROCESS_EVERY_N_FRAMES = 1  # Process every frame for better detection

# Add frame counter
frame_counter = 0

def preprocess_face(face):
    try:
        if face is None:
            return None
            
        # Ensure minimum size for better quality
        if face.shape[0] < 50 or face.shape[1] < 50:
            return None
            
        # Resize to larger size for better quality
        face = cv2.resize(face, (150, 150))
        
        # Convert to grayscale if needed
        if len(face.shape) == 3:
            face = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
        
        # Enhance contrast with more aggressive parameters
        clahe = cv2.createCLAHE(clipLimit=3.5, tileGridSize=(8,8))
        face = clahe.apply(face)
        
        # Denoise the image
        face = cv2.fastNlMeansDenoising(face)
        
        # Normalize
        face = cv2.normalize(face, None, 0, 255, cv2.NORM_MINMAX)
        
        return face
    except Exception as e:
        print(f"Error in preprocessing: {str(e)}")
        return None

def compare_faces(face1, face2):
    try:
        if face1 is None or face2 is None:
            return 0
            
        # Ensure same size
        if face1.shape != face2.shape:
            face2 = cv2.resize(face2, (face1.shape[1], face1.shape[0]))
            
        # Calculate correlation coefficient
        correlation = cv2.matchTemplate(face1, face2, cv2.TM_CCOEFF_NORMED)[0][0]
        
        # Calculate histogram similarity
        hist1 = cv2.calcHist([face1], [0], None, [256], [0,256])
        hist2 = cv2.calcHist([face2], [0], None, [256], [0,256])
        hist_sim = cv2.compareHist(hist1, hist2, cv2.HISTCMP_CORREL)
        
        # Combined similarity score
        similarity = 0.6 * correlation + 0.4 * max(0, hist_sim)
        return max(0, min(1, similarity))
        
    except Exception as e:
        print(f"Error comparing faces: {str(e)}")
        return 0

def load_known_face(image_path):
    try:
        print(f"\nLoading image: {image_path}")
        # Read image in BGR format
        img = cv2.imread(image_path)
        if img is None:
            print(f"Failed to load image: {image_path}")
            return None
            
        # Convert BGR to RGB (face_recognition requires RGB)
        rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        print(f"Image shape: {rgb_img.shape}")
        
        # Detect faces with more lenient parameters
        face_locations = face_recognition.face_locations(
            rgb_img, 
            model=MODEL,
            number_of_times_to_upsample=2  # Increase detection sensitivity
        )
        
        if len(face_locations) == 0:
            print(f"No face detected in image: {image_path}")
            return None
        
        print(f"Found {len(face_locations)} faces in image")
        
        # Get face encodings with improved parameters
        face_encodings = face_recognition.face_encodings(
            rgb_img,
            face_locations,
            num_jitters=2  # Slightly increase accuracy for known faces
        )
        
        if len(face_encodings) == 0:
            print(f"Could not encode face in: {image_path}")
            return None
            
        print(f"Successfully loaded face from: {image_path}")
        return face_encodings[0]
            
    except Exception as e:
        print(f"Error loading image {image_path}: {str(e)}")
        return None

# Get list of known faces
known_faces_names = [
    "Aman",
    "Anas",
    "Anjali",
    "Komal",
    "Kshitij",
    "Prit",
    "Rajan",
    "Siddhesh",
    "Tejas"
]

# Load known faces
print("\nLoading known faces...")
known_faces = {}
successful_loads = 0

for name in known_faces_names:
    img_path = f"{name}.jpg"
    if not os.path.exists(img_path):
        print(f"Image file not found: {img_path}")
        continue
        
    face = load_known_face(img_path)
    if face is not None:
        known_faces[name] = face
        successful_loads += 1
        print(f"Added {name} to known faces")

if successful_loads == 0:
    print("\nNo face images could be loaded. Please check the image files.")
    exit()

print(f"\nSuccessfully loaded {successful_loads} out of {len(known_faces_names)} faces")

# Modify the attendance initialization
def initialize_attendance_file():
    current_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"Attendance_{current_date}.xlsx"
    
    # Create or load workbook
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Set up headers
        ws.title = f"Attendance {current_date}"
        ws['A1'] = "Date:"
        ws['B1'] = current_date
        ws['A2'] = "Name"
        ws['B2'] = "Time"
        ws['C2'] = "Status"
        
        # Style headers
        for cell in ['A1', 'A2', 'B2', 'C2']:
            ws[cell].font = Font(bold=True)
            ws[cell].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    wb.save(filename)
    return filename

# Modify the attendance marking function
def mark_attendance(name):
    current_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"Attendance_{current_date}.xlsx"
    current_time = datetime.now().strftime("%H:%M:%S")
    
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        # Check if person is already marked
        for row in range(3, ws.max_row + 1):
            if ws[f'A{row}'].value == name:
                return False  # Already marked
        
        # Add new attendance record
        next_row = ws.max_row + 1
        ws[f'A{next_row}'] = name
        ws[f'B{next_row}'] = current_time
        ws[f'C{next_row}'] = "Present"
        
        # Style the new row
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{next_row}']
            cell.alignment = Alignment(horizontal='center')
        
        wb.save(filename)
        return True
    except Exception as e:
        print(f"Error marking attendance: {str(e)}")
        return False

# Initialize camera
print("\nInitializing camera...")
camera = None

for idx in range(2):
    print(f"Trying camera index {idx}...")
    camera = cv2.VideoCapture(idx, cv2.CAP_DSHOW)
    
    if not camera.isOpened():
        print(f"Failed to open camera {idx}")
        continue
        
    ret, frame = camera.read()
    if ret and frame is not None and frame.size > 0:
        print(f"Successfully opened camera {idx}")
        break
    else:
        print(f"Camera {idx} opened but couldn't read frame")
        camera.release()
        camera = None

if camera is None:
    print("Could not initialize any camera. Please check your camera connection.")
    exit()

# Configure camera for better performance
camera.set(cv2.CAP_PROP_FRAME_WIDTH, 640)  # Increased resolution
camera.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
camera.set(cv2.CAP_PROP_FPS, 30)
camera.set(cv2.CAP_PROP_BUFFERSIZE, 1)
camera.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'MJPG'))  # Use MJPG for better performance

print("\nCamera settings:")
print(f"Resolution: {camera.get(cv2.CAP_PROP_FRAME_WIDTH)}x{camera.get(cv2.CAP_PROP_FRAME_HEIGHT)}")
print(f"FPS: {camera.get(cv2.CAP_PROP_FPS)}")

# Modify the process_frame function
def process_frame(frame, known_faces, known_names):
    global frame_counter
    frame_counter += 1
    
    if frame_counter % PROCESS_EVERY_N_FRAMES != 0:
        return frame
    
    # Resize frame for processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
    
    # Detect faces with improved parameters
    face_locations = face_recognition.face_locations(
        rgb_small_frame, 
        model=MODEL,
        number_of_times_to_upsample=1
    )
    
    if face_locations:
        face_encodings = face_recognition.face_encodings(
            rgb_small_frame,
            face_locations,
            num_jitters=1
        )
        
        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            # Scale back up face locations
            top = int(top * 4)
            right = int(right * 4)
            bottom = int(bottom * 4)
            left = int(left * 4)
            
            # Use numpy for faster face comparison
            if len(known_faces) > 0:
                face_distances = face_recognition.face_distance(known_faces, face_encoding)
                best_match_index = np.argmin(face_distances)
                
                if face_distances[best_match_index] < TOLERANCE:
                    name = known_names[best_match_index]
                    
                    # Mark attendance and show status
                    if mark_attendance(name):
                        status_color = (0, 255, 0)  # Green for newly marked
                        print(f"✓ Marked attendance for {name}")
                    else:
                        status_color = (255, 255, 0)  # Yellow for already marked
                else:
                    name = "Unknown"
                    status_color = (0, 0, 255)  # Red for unknown
            else:
                name = "Unknown"
                status_color = (0, 0, 255)
            
            # Draw box and name with status color
            cv2.rectangle(frame, (left, top), (right, bottom), status_color, 2)
            
            # Add name label with background
            label_size = cv2.getTextSize(name, cv2.FONT_HERSHEY_SIMPLEX, 0.75, 2)[0]
            cv2.rectangle(frame, (left, top - 35), (left + label_size[0] + 10, top), status_color, cv2.FILLED)
            cv2.putText(frame, name, (left + 5, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255, 255, 255), 2)
    
    return frame

# Modify the main initialization
if __name__ == "__main__":
    # Initialize attendance file
    attendance_file = initialize_attendance_file()
    print(f"\nCreated attendance file: {attendance_file}")
    
    try:
        while True:
            ret, frame = camera.read()
            
            if not ret or frame is None or frame.size == 0:
                continue
            
            frame = process_frame(frame, list(known_faces.values()), list(known_faces.keys()))
            
            # Add date/time overlay
            current_time = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            cv2.putText(frame, current_time, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 
                       0.7, (255, 255, 255), 2)
            
            cv2.imshow('Attendance System', frame)
            
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
                
    except KeyboardInterrupt:
        print("\nStopping gracefully...")
    finally:
        camera.release()
        cv2.destroyAllWindows()
        print("\nAttendance has been saved to:", attendance_file)
        print("System shutdown successfully")
